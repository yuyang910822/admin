import openpyxl,re
# from config.path import ExcelConfig
import random,time
class DoExcel():

    def __init__(self,file_name,sheet_name):
        self.f = file_name
        self.s = sheet_name
        self.workbook = openpyxl.load_workbook(self.f)
        self.sheet = self.workbook[self.s]


    # def open(self):
    #     """打开工作薄，选择表单"""
    #     self.workbook = openpyxl.load_workbook(self.f)
    #     self.sheet = self.workbook[self.s]
    #     return self.sheet

    def close(self):
        """关闭工作薄对象，释放内存"""
        self.workbook.close()


    def firstline(self) -> list:
        """首行信息"""
        return [i.value for i in self.sheet[1]]

    def all(self) -> list:
        """测试数据"""
        # self.sheet
        testdata = []
        maxrow = self.sheet.max_row
        if  maxrow <= 1:
            print("无法拼接数据，请检查{}的{}表".format(self.f,self.s))
        elif maxrow == 2:
            ii = []
            for i in self.sheet[2:maxrow]:
                ii.append(i.value)
            z = dict(zip(self.firstline(), ii))
            testdata.append(z)
        else:
            for i in self.sheet[2:maxrow]:
                ii = []
                for j in i:
                    ii.append(j.value)
                z = dict(zip(self.firstline(), ii))
                testdata.append(z)
        self.close()
        return testdata






    def result(self,data,new,v):

        # 写入数据
        self.sheet.cell(data["id"]+1,new,v)
        # 保存文件
        self.workbook.save(ExcelConfig.testDataPant)
        # 关闭工作薄
        self.close()



    def updataData(self,data):
        """提取json进行更新"""
        name=str(random.randint(100, 9999999))
        code=str(random.randint(100, 9999999))
        data1=str(random.randint(100, 9999999))
        if name !=code:
            if re.search("name",str(data)):
                data["json"] = data["json"].replace("name",name)
            if re.search("code",str(data)):
                data["json"] = data["json"].replace("code",code)
            if re.search("data",str(data)):
                data["json"] = data["json"].replace("data",data1)


if __name__ == "__main__":
    pass

#
#
#         if not phoneid == None:
#             if re.search("phoneid",str(data)):
#                 p  = set_phoneid()
#                 data["json"] = data["json"].replace("phoneid", p)
#                 data["expect"] = data["expect"].replace("phoneid", p)
#
#         if not id == None:
#             if re.search("expectid",str(data)):
#                 if re.search("id",str(id)):
#                     i = id["data"]["id"]
#
#                     data["expect"] = data["expect"].replace("expectid", str(i))
#                     self.open()
#                     # 写入数据
#                     self.sheet.cell(data["id"] + 1, 6, data["expect"])
#                     self.sheet.cell(data["id"] + 1, 5, data["json"])
#                     # 保存文件
#                     self.workbook.save(ExcelConfig.testResuitPath)
#                     # 关闭工作薄
#                     self.close()
#
#         return data
#
#
#
if __name__ == "__main__":
    import json
    a = DoExcel(r"C:\Users\yuyang\Desktop\a.xlsx","Sheet1")
    # b = a.all()
    # for i in b:
    #     print(i)
    #
    #
    print(help(a))
